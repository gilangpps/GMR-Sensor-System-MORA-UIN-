"""
GMR Fe3O4 Sensor - Regression Model Training
Models: SVR, KNN Regressor, Random Forest Regressor
Input: Delta-B (mT) from processed sheet, all iterations
Target: Continuous concentration (mg/mL)
"""

import os
import json
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import joblib

from sklearn.pipeline import Pipeline
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split, KFold, cross_val_score
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    mean_absolute_error,
    mean_squared_error,
    r2_score,
    mean_absolute_percentage_error,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
sns.set_theme(style="whitegrid", context="talk")

# ─── CONFIG ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(_HERE)

DATA_PATH = os.path.join(REPO_ROOT, "090526_fe3o4-data-gmr.xlsx")
SHEET_NAME = "Data_Acquisition_Processed"

BASE_OUT_DIR = os.path.join(REPO_ROOT, "output_results", "regression")
MODEL_DIR = os.path.join(REPO_ROOT, "models", "regression")

os.makedirs(MODEL_DIR, exist_ok=True)

RANDOM_STATE = 42
TEST_SIZE = 0.2
CV_FOLDS = 5

CONCENTRATIONS = [5, 10, 20, 30, 40, 50]
COL_GROUPS = {
    5:  [1, 2, 3, 4, 5],
    10: [6, 7, 8, 9, 10],
    20: [11, 12, 13, 14, 15],
    30: [16, 17, 18, 19, 20],
    40: [21, 22, 23, 24, 25],
    50: [26, 27, 28, 29, 30],
}

COLORS = {
    "LinearRegression": "#1A63FF",
    "SVR": "#1A63FF",
    "KNN": "#1A63FF",
    "RandomForest": "#1A63FF",
}

# ─── UTILITIES ───────────────────────────────────────────────────────────────
def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def dirs_for_regression_model(base_out, model_name):
    """output_results/regression/<Model>/{excel,json,plots}"""
    root = os.path.join(base_out, model_name)
    return {
        "root": root,
        "excel": os.path.join(root, "excel"),
        "json": os.path.join(root, "json"),
        "plots": os.path.join(root, "plots"),
    }


def dirs_regression_comparison(base_out):
    """output_results/regression/comparison/{excel,json,plots}"""
    root = os.path.join(base_out, "comparison")
    return {
        "root": root,
        "excel": os.path.join(root, "excel"),
        "json": os.path.join(root, "json"),
        "plots": os.path.join(root, "plots"),
    }

def style_header(cell, bg="7C2D12"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def style_cell(cell):
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def load_data(path, sheet):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Data file not found: {path}\n"
            "Please place 090526_fe3o4-data-gmr.xlsx in the repository root."
        )
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    data = raw.iloc[2:].reset_index(drop=True)

    records = []
    for conc, cols in COL_GROUPS.items():
        for c in cols:
            col_data = pd.to_numeric(data.iloc[:, c], errors="coerce").dropna().values
            for val in col_data:
                records.append({"delta_B_mT": float(val), "concentration": float(conc)})

    return pd.DataFrame(records)

def build_models():
    return {
        "LinearRegression": Pipeline(
            steps=[
                ("model", LinearRegression()),
            ]
        ),
        "SVR": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", SVR(kernel="rbf", C=100, gamma="scale", epsilon=0.5)),
            ]
        ),
        "KNN": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", KNeighborsRegressor(n_neighbors=7, metric="euclidean")),
            ]
        ),
        "RandomForest": Pipeline(
            steps=[
                ("model", RandomForestRegressor(n_estimators=200, max_depth=None, random_state=RANDOM_STATE)),
            ]
        ),
    }

def _json_sanitize(obj):
    if isinstance(obj, dict):
        return {k: _json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_sanitize(v) for v in obj]
    if isinstance(obj, (np.integer, np.floating)):
        return float(obj)
    return obj


def save_metrics_json(payload, output_path):
    ensure_dir(os.path.dirname(output_path))
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(_json_sanitize(payload), f, indent=4, ensure_ascii=False)

# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────
def write_metrics_excel(all_metrics, cv_results, residual_dict, output_path):
    wb = Workbook()

    # Sheet 1: Summary Metrics
    ws = wb.active
    ws.title = "Summary Metrics"
    headers = ["Model", "MAE (mg/mL)", "MSE", "RMSE (mg/mL)", "R²", "MAPE (%)", "CV Mean R²", "CV Std R²"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for row, (name, m) in enumerate(all_metrics.items(), 2):
        vals = [name, m["mae"], m["mse"], m["rmse"], m["r2"], m["mape"], m["cv_mean"], m["cv_std"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row, col, round(v, 6) if isinstance(v, float) else v)
            style_cell(c)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Residuals per model
    for model_name, res_df in residual_dict.items():
        ws2 = wb.create_sheet(f"Residuals_{model_name}")
        for col, h in enumerate(["Actual (mg/mL)", "Predicted (mg/mL)", "Residual", "Abs Residual"], 1):
            style_header(ws2.cell(1, col, h))
        for row, (_, r) in enumerate(res_df.iterrows(), 2):
            for col, v in enumerate([r["actual"], r["predicted"], r["residual"], r["abs_residual"]], 1):
                c = ws2.cell(row, col, round(float(v), 6))
                style_cell(c)
        for col in range(1, 5):
            ws2.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 3: CV Results
    ws3 = wb.create_sheet("Cross_Validation")
    style_header(ws3.cell(1, 1, "Model"))
    for fold in range(1, CV_FOLDS + 1):
        style_header(ws3.cell(1, fold + 1, f"Fold {fold} R²"))
    style_header(ws3.cell(1, CV_FOLDS + 2, "Mean R²"))
    style_header(ws3.cell(1, CV_FOLDS + 3, "Std R²"))

    for row, (name, scores) in enumerate(cv_results.items(), 2):
        ws3.cell(row, 1, name).font = Font(bold=True, name="Arial")
        for fold, s in enumerate(scores, 2):
            c = ws3.cell(row, fold, round(float(s), 6))
            style_cell(c)
        ws3.cell(row, CV_FOLDS + 2, round(float(np.mean(scores)), 6))
        ws3.cell(row, CV_FOLDS + 3, round(float(np.std(scores)), 6))

    for col in range(1, CV_FOLDS + 4):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_path)
    print(f"  [Excel] Saved -> {output_path}")

# ─── VISUALIZATION ────────────────────────────────────────────────────────────
def plot_regression_individual(model_name, metrics, y_test, y_pred, residuals, cv_scores, plots_dir):
    ensure_dir(plots_dir)
    color = COLORS[model_name]

    # 1. Parity plot
    plt.figure(figsize=(6, 6))
    lims = [min(y_test.min(), y_pred.min()), max(y_test.max(), y_pred.max())]
    plt.scatter(y_test, y_pred, alpha=0.8, s=30, color=color, edgecolors="none")
    plt.plot(lims, lims, "k--", lw=1.2, label=f"Ideal (y=x)\nMAE={metrics['mae']:.3f}, RMSE={metrics['rmse']:.3f}, R²={metrics['r2']:.3f}")
    for xi, yi in zip(y_test, y_pred):
        plt.text(xi, yi, f"{yi:.2f}", fontsize=7, alpha=0.6)
    plt.xlabel("Actual Concentration (mg/mL)")
    plt.ylabel("Predicted Concentration (mg/mL)")
    plt.title(f"Parity Plot — {model_name}")
    plt.legend(fontsize=8)
    plt.grid(True, ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "parity.png"), dpi=200)
    plt.close()

    # 2. Residual plot
    plt.figure(figsize=(7, 5))
    plt.scatter(y_pred, residuals, alpha=0.8, s=30, color=color, edgecolors="none")
    for xi, yi in zip(y_pred, residuals):
        plt.text(xi, yi, f"{yi:.2f}", fontsize=7, alpha=0.6)
    plt.axhline(0, color="k", lw=1)
    plt.xlabel("Predicted Concentration (mg/mL)")
    plt.ylabel("Residual (mg/mL)")
    plt.title(f"Residuals — {model_name}")
    plt.grid(True, ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "residuals.png"), dpi=200)
    plt.close()

    # 3. Residual distribution
    plt.figure(figsize=(7, 5.5))
    plt.hist(residuals, bins=30, color=color, alpha=0.75, edgecolor="white")
    plt.axvline(0, color="k", lw=1)
    plt.xlabel("Residual (mg/mL)")
    plt.ylabel("Count")
    plt.title(f"Residual Distribution — {model_name}")
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "residual_distribution.png"), dpi=200)
    plt.close()

    # 4. Overall metrics summary
    plt.figure(figsize=(8, 5))
    metric_names = ["MAE", "RMSE", "R²", "MAPE (%)"]
    metric_vals = [metrics["mae"], metrics["rmse"], metrics["r2"], metrics["mape"]]
    bars = plt.bar(metric_names, metric_vals, color=color, alpha=0.85, edgecolor="white")
    plt.title(f"Overall Metrics — {model_name}")
    plt.ylabel("Value")
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    mv_max = max(metric_vals) if metric_vals else 1.0
    for bar, v in zip(bars, metric_vals):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + mv_max * 0.01,
            f"{v:.4f}",
            ha="center",
            va="bottom",
            fontsize=8,
        )
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "overall_metrics.png"), dpi=200)
    plt.close()

    # 5. Absolute error per concentration class
    plt.figure(figsize=(9, 5.5))
    groups = {c: [] for c in CONCENTRATIONS}
    for actual, pred in zip(y_test, y_pred):
        c_key = int(round(actual))
        if c_key in groups:
            groups[c_key].append(abs(actual - pred))
    box_data = [groups[c] for c in CONCENTRATIONS]
    bp = plt.boxplot(box_data, patch_artist=True, labels=[str(c) for c in CONCENTRATIONS])
    for patch in bp["boxes"]:
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    plt.xlabel("Concentration (mg/mL)")
    plt.ylabel("Absolute Error (mg/mL)")
    plt.title(f"Error per Concentration — {model_name}")
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "error_per_concentration.png"), dpi=200)
    plt.close()

    # 6. CV R² across folds
    plt.figure(figsize=(8, 5))
    plt.bar([f"Fold {i+1}" for i in range(len(cv_scores))], cv_scores, color=color, alpha=0.85, edgecolor="white")
    plt.axhline(np.mean(cv_scores), color="black", ls="--", lw=1.2, label=f"Mean = {np.mean(cv_scores):.4f}")
    plt.title(f"{CV_FOLDS}-Fold CV R² — {model_name}")
    plt.ylabel("R²")
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    plt.legend(fontsize=8)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "cv_r2_folds.png"), dpi=200)
    plt.close()

    print(f"  [Plots] Saved -> {plots_dir} (6 figures: parity, residuals, metrics, ...)")

def plot_regression_comparison(all_metrics, cv_results, y_test, preds, output_dir):
    ensure_dir(output_dir)

    fig = plt.figure(figsize=(22, 14))
    fig.suptitle(
        "Regression Comparison — SVR vs KNN vs Random Forest\nGMR Fe₃O₄ Sensor (ΔB, mT)",
        fontsize=15,
        fontweight="bold",
        y=0.99,
    )
    gs = gridspec.GridSpec(3, 4, figure=fig, hspace=0.55, wspace=0.4)

    model_names = list(all_metrics.keys())
    color_map = [COLORS[n] for n in model_names]

    # Row 0: Parity plots
    for idx, (name, color) in enumerate(zip(model_names, color_map)):
        ax = fig.add_subplot(gs[0, idx])
        y_pred = preds[name]
        ax.scatter(y_test, y_pred, alpha=0.6, s=20, color=color, edgecolors="none")
        lims = [min(y_test.min(), y_pred.min()), max(y_test.max(), y_pred.max())]
        ax.plot(lims, lims, "k--", lw=1.2)
        ax.set_xlabel("Actual (mg/mL)", fontsize=8)
        ax.set_ylabel("Predicted (mg/mL)", fontsize=8)
        ax.set_title(f"{name} — Parity (R² = {all_metrics[name]['r2']:.4f})", fontweight="bold", fontsize=10)
        ax.grid(True, ls="--", alpha=0.4)

    # Row 1 left: Metric bar comparison
    ax_bar = fig.add_subplot(gs[1, :3])
    metric_keys = ["mae", "rmse", "r2", "mape"]
    metric_labels = ["MAE (mg/mL)", "RMSE (mg/mL)", "R²", "MAPE (%)"]
    x = np.arange(len(metric_labels))
    width = 0.25
    for i, (name, color) in enumerate(zip(model_names, color_map)):
        vals = [all_metrics[name][k] for k in metric_keys]
        bars = ax_bar.bar(x + (i - 1.5) * width, vals, width, label=name, color=color, alpha=0.85, edgecolor="white")
        for bar, v in zip(bars, vals):
            ax_bar.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(vals) * 0.01,
                f"{v:.3f}",
                ha="center",
                va="bottom",
                fontsize=8,
            )
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(metric_labels)
    ax_bar.set_title("Metric Comparison (All Models)", fontweight="bold")
    ax_bar.set_ylabel("Value")
    ax_bar.grid(True, axis="y", ls="--", alpha=0.4)
    ax_bar.legend()

    # Row 1 right: CV R² boxplot
    ax_cv = fig.add_subplot(gs[1, 3])
    cv_data = [cv_results[n] for n in model_names]
    bp = ax_cv.boxplot(cv_data, patch_artist=True, labels=model_names)
    for patch, color in zip(bp["boxes"], color_map):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    ax_cv.set_title(f"{CV_FOLDS}-Fold CV R²", fontweight="bold")
    ax_cv.set_ylabel("R²")
    ax_cv.grid(True, axis="y", ls="--", alpha=0.4)

    # Row 2: Residual distributions
    for idx, (name, color) in enumerate(zip(model_names, color_map)):
        ax = fig.add_subplot(gs[2, idx])
        res = y_test - preds[name]
        ax.hist(res, bins=30, color=color, alpha=0.75, edgecolor="white")
        ax.axvline(0, color="black", lw=1.5, ls="--")
        ax.set_xlabel("Residual (mg/mL)", fontsize=8)
        ax.set_ylabel("Count", fontsize=8)
        ax.set_title(f"{name} — Residual Distribution", fontweight="bold", fontsize=10)
        ax.grid(True, axis="y", ls="--", alpha=0.4)

    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_path = os.path.join(output_dir, "regression_comparison_all_models.png")
    plt.savefig(out_path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  [Plot] Saved -> {out_path}")

def plot_regression_metrics_comparison(all_metrics, output_dir):
    ensure_dir(output_dir)
    
    model_names = list(all_metrics.keys())
    color_map = [COLORS[n] for n in model_names]
    
    metrics = [
        ("mae", "MAE (mg/mL)", "mae_comparison.png"),
        ("rmse", "RMSE (mg/mL)", "rmse_comparison.png"),
        ("r2", "R²", "r2_comparison.png"),
        ("mape", "MAPE (%)", "mape_comparison.png"),
    ]
    
    for key, label, filename in metrics:
        plt.figure(figsize=(8, 6))
        vals = [all_metrics[name][key] for name in model_names]
        bars = plt.bar(model_names, vals, color=color_map, alpha=0.85, edgecolor="white")
        plt.title(f"Regression {label} Comparison — KNN vs SVR vs Random Forest", fontweight="bold")
        plt.ylabel(label)
        plt.grid(True, axis="y", ls="--", alpha=0.4)
        for bar, v in zip(bars, vals):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(vals) * 0.01,
                f"{v:.4f}",
                ha="center",
                va="bottom",
                fontsize=10,
            )
        plt.tight_layout()
        out_path = os.path.join(output_dir, filename)
        plt.savefig(out_path, dpi=200, bbox_inches="tight")
        plt.close()
        print(f"  [Plot] Saved -> {out_path}")


def plot_best_model_parity(all_metrics, y_test, preds, output_dir):
    ensure_dir(output_dir)
    best_model_name = min(all_metrics, key=lambda n: all_metrics[n]["rmse"])
    best_metrics = all_metrics[best_model_name]
    y_pred = preds[best_model_name]

    friendly_names = {
        "LinearRegression": "Linear Regression",
        "SVR": "SVR",
        "KNN": "KNN Regression",
        "RandomForest": "Random Forest Regression",
    }
    title_name = friendly_names.get(best_model_name, best_model_name)

    fig, ax = plt.subplots(figsize=(8, 6))
    lims = [min(y_test.min(), y_pred.min()), max(y_test.max(), y_pred.max())]
    ax.scatter(y_test, y_pred, alpha=0.8, s=40, color=COLORS.get(best_model_name, "#1f77b4"), edgecolors="none")
    ax.plot(lims, lims, linestyle="--", color="#2c7fb8", linewidth=2, label="Ideal prediction")
    ax.set_xlim(lims)
    ax.set_ylim(lims)
    ax.set_xlabel("Actual Fe₃O₄ concentration (mg/mL)")
    ax.set_ylabel("Predicted Fe₃O₄ concentration (mg/mL)")
    ax.set_title("Regression-Based Concentration Prediction", fontsize=16, fontweight="bold")
    ax.grid(True, linestyle="--", alpha=0.4)

    textstr = (
        f"Best model: {title_name}\n"
        f"MAE = {best_metrics['mae']:.3f} mg/mL\n"
        f"RMSE = {best_metrics['rmse']:.3f} mg/mL\n"
        f"R² = {best_metrics['r2']:.3f}"
    )
    props = dict(boxstyle="round", facecolor="white", edgecolor="black", alpha=0.85)
    ax.text(0.03, 0.97, textstr, transform=ax.transAxes, fontsize=10, va="top", bbox=props)
    ax.legend(loc="lower right", fontsize=9)

    plt.tight_layout()
    out_path = os.path.join(output_dir, "best_model_parity.png")
    plt.savefig(out_path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  [Plot] Saved -> {out_path}")


def plot_multi_metric_comparison(all_metrics, model_names, color_map, output_dir, filename, title):
    metrics = ["mae", "rmse", "r2", "mape"]
    labels = ["MAE", "RMSE", "R²", "MAPE"]
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    axes = axes.flatten()
    for i, (metric, label) in enumerate(zip(metrics, labels)):
        ax = axes[i]
        vals = [all_metrics[name][metric] for name in model_names]
        bars = ax.bar(model_names, vals, color=color_map[:len(model_names)], alpha=0.85, edgecolor="white")
        ax.set_title(f"{label} Comparison", fontweight="bold")
        ax.set_ylabel(label)
        ax.grid(True, axis="y", ls="--", alpha=0.4)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(vals)*0.01, f"{v:.3f}", ha="center", va="bottom", fontsize=8)
        ax.set_xticklabels(model_names, rotation=45, ha="right")
    plt.suptitle(title, fontsize=16, fontweight="bold")
    plt.tight_layout()
    out_path = os.path.join(output_dir, filename)
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"  [Multiplot] Saved -> {out_path}")


def plot_multi_metric_comparison_separate(all_metrics, model_names, color_map, output_dir, prefix, title):
    metrics = ["mae", "rmse", "r2", "mape"]
    labels = ["MAE", "RMSE", "R²", "MAPE"]
    for metric, label in zip(metrics, labels):
        plt.figure(figsize=(8, 6))
        vals = [all_metrics[name][metric] for name in model_names]
        bars = plt.bar(model_names, vals, color=color_map[:len(model_names)], alpha=0.85, edgecolor="white")
        plt.title(f"{title} — {label}", fontweight="bold")
        plt.ylabel(label)
        plt.grid(True, axis="y", ls="--", alpha=0.4)
        if metric in ["r2"]:
            plt.ylim(min(0, min(vals) - 0.05), max(vals) + max(vals) * 0.05)
        for bar, v in zip(bars, vals):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(vals) * 0.01, f"{v:.3f}", ha="center", va="bottom", fontsize=8)
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        out_path = os.path.join(output_dir, f"{prefix}_{metric}.png")
        plt.savefig(out_path, dpi=300, bbox_inches="tight")
        plt.close()
        print(f"  [Multiplot] Saved -> {out_path}")

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  GMR Fe3O4 — Regression Training")
    print("=" * 60)

    df = load_data(DATA_PATH, SHEET_NAME)
    print(f"  Dataset: {len(df)} samples")

    X = df[["delta_B_mT"]].values
    y = df["concentration"].values

    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=TEST_SIZE,
        random_state=RANDOM_STATE,
    )
    print(f"  Train: {len(X_train)} | Test: {len(X_test)}")

    models = build_models()
    kf = KFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)

    all_metrics = {}
    cv_results = {}
    residual_dict = {}
    preds = {}

    for name, model in models.items():
        print(f"\n  Training {name}...")
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)

        mae = mean_absolute_error(y_test, y_pred)
        mse = mean_squared_error(y_test, y_pred)
        rmse = np.sqrt(mse)
        r2 = r2_score(y_test, y_pred)
        mape = mean_absolute_percentage_error(y_test, y_pred) * 100
        cv_sc = cross_val_score(model, X, y, cv=kf, scoring="r2")

        all_metrics[name] = {
            "mae": float(mae),
            "mse": float(mse),
            "rmse": float(rmse),
            "r2": float(r2),
            "mape": float(mape),
            "cv_mean": float(np.mean(cv_sc)),
            "cv_std": float(np.std(cv_sc)),
        }
        cv_results[name] = cv_sc
        preds[name] = y_pred

        res = y_test - y_pred
        residual_dict[name] = pd.DataFrame(
            {
                "actual": y_test,
                "predicted": y_pred,
                "residual": res,
                "abs_residual": np.abs(res),
            }
        )

        print(f"    MAE  : {mae:.4f} mg/mL")
        print(f"    RMSE : {rmse:.4f} mg/mL")
        print(f"    R²   : {r2:.4f}")
        print(f"    MAPE : {mape:.2f} %")
        print(f"    CV R²: {np.mean(cv_sc):.4f} ± {np.std(cv_sc):.4f}")

        model_dir = os.path.join(MODEL_DIR, name)
        ensure_dir(model_dir)
        model_path = os.path.join(model_dir, f"regressor_{name.lower()}.pkl")
        joblib.dump(model, model_path)
        print(f"    [Model] Saved -> {model_path}")

        out = dirs_for_regression_model(BASE_OUT_DIR, name)
        for key in ("excel", "json", "plots"):
            ensure_dir(out[key])

        plot_regression_individual(
            model_name=name,
            metrics=all_metrics[name],
            y_test=y_test,
            y_pred=y_pred,
            residuals=res,
            cv_scores=cv_sc,
            plots_dir=out["plots"],
        )

        write_metrics_excel(
            {name: all_metrics[name]},
            {name: cv_results[name]},
            {name: residual_dict[name]},
            os.path.join(out["excel"], "regression_metrics.xlsx"),
        )
        save_metrics_json(
            {"model": name, "metrics": all_metrics[name]},
            os.path.join(out["json"], "regression_metrics.json"),
        )
        print(f"    [JSON] Saved -> {out['json']}")

    comp = dirs_regression_comparison(BASE_OUT_DIR)
    for key in ("excel", "json", "plots"):
        ensure_dir(comp[key])

    excel_path = os.path.join(comp["excel"], "regression_metrics.xlsx")
    write_metrics_excel(all_metrics, cv_results, residual_dict, excel_path)

    json_path = os.path.join(comp["json"], "regression_metrics.json")
    save_metrics_json(all_metrics, json_path)
    print(f"  [JSON] Saved -> {json_path}")

    print("\n  Generating comparison plot...")
    plot_regression_comparison(
        all_metrics=all_metrics,
        cv_results=cv_results,
        y_test=y_test,
        preds=preds,
        output_dir=comp["plots"],
    )

    print("\n  Generating best-model parity plot...")
    plot_best_model_parity(all_metrics, y_test, preds, comp["plots"])

    print("\n  Generating metrics comparison plot...")
    plot_regression_metrics_comparison(all_metrics, comp["plots"])

    print("\n  Generating multi-metric comparison plots...")
    multiplot_root = os.path.join(comp["plots"], "multiplots")
    multiplot_4_dir = os.path.join(multiplot_root, "4_models")
    multiplot_3_dir = os.path.join(multiplot_root, "3_models")
    os.makedirs(multiplot_4_dir, exist_ok=True)
    os.makedirs(multiplot_3_dir, exist_ok=True)

    # 4 models multiplot
    plot_multi_metric_comparison(
        all_metrics,
        list(all_metrics.keys()),
        [COLORS[n] for n in all_metrics.keys()],
        multiplot_4_dir,
        "multiplot_4_models.png",
        "Regression Models Comparison (4 Models)"
    )

    # 3 models separate per-metric plots (SVR, KNN, RandomForest)
    three_models = ["SVR", "KNN", "RandomForest"]
    three_metrics = {k: v for k, v in all_metrics.items() if k in three_models}
    three_colors = [COLORS[n] for n in three_models]
    plot_multi_metric_comparison_separate(
        three_metrics,
        three_models,
        three_colors,
        multiplot_3_dir,
        "multiplot_3_models",
        "Regression Models Comparison (SVR, KNN, RF)"
    )

    print("\n" + "=" * 60)
    print("  Regression training complete.")
    print("=" * 60)

if __name__ == "__main__":
    main()