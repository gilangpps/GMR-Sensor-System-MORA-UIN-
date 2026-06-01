"""
GMR Fe3O4 Sensor - Classification Model Training
Models: SVC, KNN, Random Forest
Input: Delta-B (mT) from processed sheet, all iterations
Target: Concentration class (5, 10, 20, 30, 40, 50 mg/mL)
"""

import os
import json
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import joblib

from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    confusion_matrix,
    classification_report,
    roc_auc_score,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
sns.set_theme(style="whitegrid", context="talk")

# ─── CONFIG ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(_HERE)

DATA_PATH = os.path.join(REPO_ROOT, "090526_fe3o4-data-gmr.xlsx")
SHEET_NAME = "Data_Acquisition_Processed"

BASE_OUT_DIR = os.path.join(REPO_ROOT, "output_results", "classification")
MODEL_DIR = os.path.join(REPO_ROOT, "models", "classification")

os.makedirs(MODEL_DIR, exist_ok=True)

RANDOM_STATE = 42
TEST_SIZE = 0.2
CV_FOLDS = 5

CONCENTRATIONS = [5, 10, 20, 30, 40, 50]
CLASS_NAMES = [f"{c} mg/mL" for c in CONCENTRATIONS]

COL_GROUPS = {
    5:  [1, 2, 3, 4, 5],
    10: [6, 7, 8, 9, 10],
    20: [11, 12, 13, 14, 15],
    30: [16, 17, 18, 19, 20],
    40: [21, 22, 23, 24, 25],
    50: [26, 27, 28, 29, 30],
}

COLORS = {
    "LogisticRegression": "#1A63FF",
    "SVM": "#1A63FF",
    "KNN": "#1A63FF",
    "RandomForest": "#1A63FF",
}

# ─── UTILITIES ───────────────────────────────────────────────────────────────
def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def dirs_for_classification_model(base_out, model_name):
    """output_results/classification/<Model>/{excel,json,plots}"""
    root = os.path.join(base_out, model_name)
    return {
        "root": root,
        "excel": os.path.join(root, "excel"),
        "json": os.path.join(root, "json"),
        "plots": os.path.join(root, "plots"),
    }


def dirs_classification_comparison(base_out):
    """output_results/classification/comparison/{excel,json,plots}"""
    root = os.path.join(base_out, "comparison")
    return {
        "root": root,
        "excel": os.path.join(root, "excel"),
        "json": os.path.join(root, "json"),
        "plots": os.path.join(root, "plots"),
    }

def style_header(cell, bg="1E40AF"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def style_cell(cell):
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def load_data(path, sheet):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Data file not found: {path}\n"
            "Please place 090526_fe3o4-data-gmr.xlsx in the repository root."
        )
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    data = raw.iloc[2:].reset_index(drop=True)

    records = []
    for conc, cols in COL_GROUPS.items():
        for c in cols:
            col_data = pd.to_numeric(data.iloc[:, c], errors="coerce").dropna().values
            for val in col_data:
                records.append({"delta_B_mT": float(val), "concentration": int(conc)})

    return pd.DataFrame(records)

def build_models():
    return {
        "LogisticRegression": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", LogisticRegression(random_state=RANDOM_STATE, max_iter=1000)),
            ]
        ),
        "SVM": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", SVC(kernel="rbf", C=10, gamma="scale", probability=True, random_state=RANDOM_STATE)),
            ]
        ),
        "KNN": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", KNeighborsClassifier(n_neighbors=7, metric="euclidean")),
            ]
        ),
        "RandomForest": Pipeline(
            steps=[
                ("model", RandomForestClassifier(n_estimators=200, max_depth=None, random_state=RANDOM_STATE)),
            ]
        ),
    }

def _json_sanitize(obj):
    if isinstance(obj, dict):
        return {k: _json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_sanitize(v) for v in obj]
    if isinstance(obj, (np.integer, np.floating)):
        return float(obj)
    return obj


def save_metrics_json(payload, output_path):
    ensure_dir(os.path.dirname(output_path))
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(_json_sanitize(payload), f, indent=4, ensure_ascii=False)

# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────
def write_metrics_excel(all_metrics, cv_results, report_dicts, output_path):
    wb = Workbook()

    # Sheet 1: Summary Metrics
    ws = wb.active
    ws.title = "Summary Metrics"
    headers = [
        "Model",
        "Accuracy",
        "Precision (macro)",
        "Recall (macro)",
        "F1 (macro)",
        "ROC-AUC (OvR)",
        "CV Mean Acc",
        "CV Std Acc",
    ]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for row, (name, m) in enumerate(all_metrics.items(), 2):
        vals = [
            name,
            m["accuracy"],
            m["precision"],
            m["recall"],
            m["f1"],
            m["roc_auc"],
            m["cv_mean"],
            m["cv_std"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row, col, round(v, 6) if isinstance(v, float) else v)
            style_cell(c)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 2: Per-Class Reports
    for model_name, report in report_dicts.items():
        ws2 = wb.create_sheet(f"Report_{model_name}")
        header_row = ["Class", "Precision", "Recall", "F1-Score", "Support"]
        for col, h in enumerate(header_row, 1):
            style_header(ws2.cell(1, col, h))

        for row, cls_name in enumerate(CLASS_NAMES, 2):
            m = report.get(cls_name, {})
            vals = [
                cls_name,
                m.get("precision", ""),
                m.get("recall", ""),
                m.get("f1-score", ""),
                m.get("support", ""),
            ]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row, col, round(v, 6) if isinstance(v, float) else v)
                style_cell(c)

        for col in range(1, 6):
            ws2.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 3: CV Results
    ws3 = wb.create_sheet("Cross_Validation")
    style_header(ws3.cell(1, 1, "Model"))
    for fold in range(1, CV_FOLDS + 1):
        style_header(ws3.cell(1, fold + 1, f"Fold {fold}"))
    style_header(ws3.cell(1, CV_FOLDS + 2, "Mean"))
    style_header(ws3.cell(1, CV_FOLDS + 3, "Std"))

    for row, (name, scores) in enumerate(cv_results.items(), 2):
        ws3.cell(row, 1, name).font = Font(bold=True, name="Arial")
        for fold, s in enumerate(scores, 2):
            c = ws3.cell(row, fold, round(float(s), 6))
            style_cell(c)
        ws3.cell(row, CV_FOLDS + 2, round(float(np.mean(scores)), 6))
        ws3.cell(row, CV_FOLDS + 3, round(float(np.std(scores)), 6))

    for col in range(1, CV_FOLDS + 4):
        ws3.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    print(f"  [Excel] Saved -> {output_path}")

# ─── VISUALIZATION ────────────────────────────────────────────────────────────
def plot_classification_individual(model_name, metrics, cm, y_test, y_pred, report, plots_dir):
    ensure_dir(plots_dir)
    color = COLORS[model_name]

    # 1. Confusion matrix
    plt.figure(figsize=(8, 6.5))
    sns.heatmap(
        cm,
        annot=True,
        fmt="d",
        cmap=sns.light_palette(color, as_cmap=True),
        linewidths=0.5,
        annot_kws={"size": 8},
        xticklabels=CLASS_NAMES,
        yticklabels=CLASS_NAMES,
    )
    plt.title(f"{model_name} — Confusion Matrix", fontweight="bold")
    plt.xlabel("Predicted")
    plt.ylabel("Actual")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "confusion_matrix.png"), dpi=200)
    plt.close()

    # 2. Per-class F1
    plt.figure(figsize=(10, 5.5))
    f1s = [report[c]["f1-score"] for c in CLASS_NAMES]
    bars = plt.bar(CLASS_NAMES, f1s, color=color, alpha=0.85, edgecolor="white", linewidth=0.7)
    plt.title(f"{model_name} — Per-Class F1 Score", fontweight="bold")
    plt.ylim(0, 1.05)
    plt.ylabel("F1 Score")
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    for bar, v in zip(bars, f1s):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.02,
            f"{v:.3f}",
            ha="center",
            va="bottom",
            fontsize=8,
        )
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "f1_score.png"), dpi=200)
    plt.close()

    # 3. Precision & Recall per class
    plt.figure(figsize=(10, 5.5))
    prec = [report[c]["precision"] for c in CLASS_NAMES]
    rec = [report[c]["recall"] for c in CLASS_NAMES]
    x = np.arange(len(CLASS_NAMES))
    plt.bar(x - 0.2, prec, 0.35, label="Precision", color=color, alpha=0.85, edgecolor="white")
    plt.bar(x + 0.2, rec, 0.35, label="Recall", color="#7C3AED", alpha=0.85, edgecolor="white")
    plt.title(f"{model_name} — Precision & Recall", fontweight="bold")
    plt.xticks(x, CLASS_NAMES, rotation=45)
    plt.ylim(0, 1.1)
    plt.ylabel("Score")
    plt.legend()
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "precision_recall.png"), dpi=200)
    plt.close()

    # 4. Overall metrics summary
    plt.figure(figsize=(9, 5.5))
    metric_names = ["Accuracy", "Precision\n(macro)", "Recall\n(macro)", "F1\n(macro)", "ROC-AUC"]
    metric_vals = [metrics["accuracy"], metrics["precision"], metrics["recall"], metrics["f1"], metrics["roc_auc"]]
    bars2 = plt.bar(metric_names, metric_vals, color=color, alpha=0.85, edgecolor="white")
    plt.ylim(0, 1.15)
    plt.title(f"{model_name} — Overall Metrics", fontweight="bold")
    plt.ylabel("Score")
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    for bar, v in zip(bars2, metric_vals):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.02,
            f"{v:.4f}",
            ha="center",
            va="bottom",
            fontsize=8,
        )
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "overall_metrics.png"), dpi=200)
    plt.close()

    # 5. Actual vs Predicted
    plt.figure(figsize=(7, 6))
    rng = np.random.default_rng(RANDOM_STATE)
    jitter = rng.uniform(-0.3, 0.3, len(y_test))
    colors_pt = [color if p == a else "#F87171" for p, a in zip(y_pred, y_test)]
    plt.scatter(y_test + jitter, y_pred + jitter, c=colors_pt, alpha=0.65, s=20)
    plt.plot([5, 50], [5, 50], "k--", lw=1.2, label=f"Ideal\nAcc={metrics['accuracy']:.3f}, Prec={metrics['precision']:.3f}, Rec={metrics['recall']:.3f}, F1={metrics['f1']:.3f}")
    plt.xlabel("Actual Concentration (mg/mL)")
    plt.ylabel("Predicted Concentration (mg/mL)")
    plt.title(f"{model_name} — Actual vs Predicted")
    plt.xticks(CONCENTRATIONS)
    plt.yticks(CONCENTRATIONS)
    plt.legend(fontsize=8)
    plt.grid(True, ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "actual_vs_predicted.png"), dpi=200)
    plt.close()

    # 6. Sample counts
    plt.figure(figsize=(10, 5.5))
    counts_actual = [np.sum(y_test == c) for c in CONCENTRATIONS]
    counts_pred = [np.sum(y_pred == c) for c in CONCENTRATIONS]
    xb = np.arange(len(CLASS_NAMES))
    plt.bar(xb - 0.2, counts_actual, 0.35, label="Actual", color="#64748B", alpha=0.85, edgecolor="white")
    plt.bar(xb + 0.2, counts_pred, 0.35, label="Predicted", color=color, alpha=0.85, edgecolor="white")
    plt.xticks(xb, CLASS_NAMES, rotation=45)
    plt.title(f"{model_name} — Sample Count: Actual vs Predicted", fontweight="bold")
    plt.ylabel("Count")
    plt.legend()
    plt.grid(True, axis="y", ls="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "sample_count_comparison.png"), dpi=200)
    plt.close()

    print(f"  [Plots] Saved -> {plots_dir} (6 figures: confusion matrix, f1, precision/recall, metrics, actual vs predicted, sample counts)")

def plot_classification_comparison(all_metrics, cv_results, report_dicts, cm_dict, y_test, preds, output_dir):
    ensure_dir(output_dir)

    fig = plt.figure(figsize=(22, 14))
    fig.suptitle(
        "Classification Comparison — SVM vs KNN vs Random Forest\nGMR Fe₃O₄ Sensor (ΔB, mT)",
        fontsize=15,
        fontweight="bold",
        y=0.99,
    )
    gs = gridspec.GridSpec(3, 4, figure=fig, hspace=0.55, wspace=0.4)

    model_names = list(all_metrics.keys())
    color_map = [COLORS[n] for n in model_names]
    metric_keys = ["accuracy", "precision", "recall", "f1", "roc_auc"]
    metric_labels = ["Accuracy", "Precision", "Recall", "F1", "ROC-AUC"]

    # Row 0: Confusion Matrices
    for idx, (name, color) in enumerate(zip(model_names, color_map)):
        ax = fig.add_subplot(gs[0, idx])
        sns.heatmap(
            cm_dict[name],
            annot=True,
            fmt="d",
            cmap=sns.light_palette(color, as_cmap=True),
            ax=ax,
            xticklabels=CLASS_NAMES,
            yticklabels=CLASS_NAMES,
            linewidths=0.5,
            annot_kws={"size": 7},
        )
        ax.set_title(f"{name} — Confusion Matrix", fontweight="bold", fontsize=10)
        ax.set_xlabel("Predicted", fontsize=8)
        ax.set_ylabel("Actual", fontsize=8)
        ax.tick_params(axis="x", rotation=45, labelsize=7)
        ax.tick_params(axis="y", rotation=0, labelsize=7)

    # Row 1 left: Grouped metric comparison
    ax_bar = fig.add_subplot(gs[1, :3])
    x = np.arange(len(metric_labels))
    width = 0.2
    for i, (name, color) in enumerate(zip(model_names, color_map)):
        vals = [all_metrics[name][k] for k in metric_keys]
        bars = ax_bar.bar(x + (i - 1.5) * width, vals, width, label=name, color=color, alpha=0.85, edgecolor="white")
        for bar, v in zip(bars, vals):
            ax_bar.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(vals) * 0.01,
                f"{v:.3f}",
                ha="center",
                va="bottom",
                fontsize=8,
            )
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(metric_labels)
    ax_bar.set_ylim(0, 1.15)
    ax_bar.set_ylabel("Score")
    ax_bar.set_title("Overall Metrics Comparison", fontweight="bold")
    ax_bar.grid(True, axis="y", ls="--", alpha=0.4)
    ax_bar.legend()

    # Row 1 right: CV boxplot
    ax_cv = fig.add_subplot(gs[1, 3])
    cv_data = [cv_results[n] for n in model_names]
    bp = ax_cv.boxplot(cv_data, patch_artist=True, labels=model_names)
    for patch, color in zip(bp["boxes"], color_map):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    ax_cv.set_title(f"{CV_FOLDS}-Fold CV Accuracy", fontweight="bold")
    ax_cv.set_ylabel("Accuracy")
    ax_cv.grid(True, axis="y", ls="--", alpha=0.4)

    # Row 2: Per-class F1 for each model
    for idx, (name, color) in enumerate(zip(model_names, color_map)):
        ax = fig.add_subplot(gs[2, idx])
        report = report_dicts[name]
        f1s = [report[c]["f1-score"] for c in CLASS_NAMES]
        bars = ax.bar(CLASS_NAMES, f1s, color=color, alpha=0.85, edgecolor="white")
        ax.set_ylim(0, 1.2)
        ax.set_ylabel("F1 Score", fontsize=8)
        ax.set_title(f"{name} — Per-Class F1", fontweight="bold", fontsize=10)
        ax.tick_params(axis="x", rotation=45, labelsize=7)
        ax.grid(True, axis="y", ls="--", alpha=0.4)
        for bar, v in zip(bars, f1s):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.02,
                f"{v:.3f}",
                ha="center",
                va="bottom",
                fontsize=7,
            )

    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_path = os.path.join(output_dir, "classification_comparison_all_models.png")
    plt.savefig(out_path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  [Plot] Saved -> {out_path}")

def plot_classification_metrics_comparison(all_metrics, output_dir):
    ensure_dir(output_dir)
    
    model_names = list(all_metrics.keys())
    color_map = [COLORS[n] for n in model_names]
    
    metrics = [
        ("accuracy", "Accuracy", "accuracy_comparison.png"),
        ("precision", "Precision (macro)", "precision_comparison.png"),
        ("recall", "Recall (macro)", "recall_comparison.png"),
        ("f1", "F1 Score (macro)", "f1_comparison.png"),
    ]
    
    for key, label, filename in metrics:
        plt.figure(figsize=(8, 6))
        vals = [all_metrics[name][key] for name in model_names]
        bars = plt.bar(model_names, vals, color=color_map, alpha=0.85, edgecolor="white")
        plt.title(f"Classification {label} Comparison — SVM vs KNN vs Random Forest", fontweight="bold")
        plt.ylabel(label)
        plt.ylim(0, 1.05)
        plt.grid(True, axis="y", ls="--", alpha=0.4)
        for bar, v in zip(bars, vals):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.01,
                f"{v:.4f}",
                ha="center",
                va="bottom",
                fontsize=10,
            )
        plt.tight_layout()
        out_path = os.path.join(output_dir, filename)
        plt.savefig(out_path, dpi=200, bbox_inches="tight")
        plt.close()
        print(f"  [Plot] Saved -> {out_path}")


def plot_multi_metric_comparison_classification(all_metrics, model_names, color_map, output_dir, filename, title):
    metrics = ["accuracy", "precision", "recall", "f1", "roc_auc"]
    labels = ["Accuracy", "Precision", "Recall", "F1", "ROC-AUC"]
    fig, axes = plt.subplots(2, 3, figsize=(15, 10))
    axes = axes.flatten()
    for i, (metric, label) in enumerate(zip(metrics, labels)):
        ax = axes[i]
        vals = [all_metrics[name][metric] for name in model_names]
        bars = ax.bar(model_names, vals, color=color_map[:len(model_names)], alpha=0.85, edgecolor="white")
        ax.set_title(f"{label} Comparison", fontweight="bold")
        ax.set_ylabel(label)
        ax.grid(True, axis="y", ls="--", alpha=0.4)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(vals)*0.01, f"{v:.3f}", ha="center", va="bottom", fontsize=8)
        ax.set_xticklabels(model_names, rotation=45, ha="right")
    # Hide the last subplot if 5 metrics
    axes[5].set_visible(False)
    plt.suptitle(title, fontsize=16, fontweight="bold")
    plt.tight_layout()
    out_path = os.path.join(output_dir, filename)
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"  [Multiplot] Saved -> {out_path}")


def plot_multi_metric_comparison_classification_separate(all_metrics, model_names, color_map, output_dir, prefix, title):
    metrics = ["accuracy", "precision", "recall", "f1", "roc_auc"]
    labels = ["Accuracy", "Precision", "Recall", "F1", "ROC-AUC"]
    for metric, label in zip(metrics, labels):
        plt.figure(figsize=(8, 6))
        vals = [all_metrics[name][metric] for name in model_names]
        bars = plt.bar(model_names, vals, color=color_map[:len(model_names)], alpha=0.85, edgecolor="white")
        plt.title(f"{title} — {label}", fontweight="bold")
        plt.ylabel(label)
        plt.ylim(0, 1.05)
        plt.grid(True, axis="y", ls="--", alpha=0.4)
        for bar, v in zip(bars, vals):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.01, f"{v:.4f}", ha="center", va="bottom", fontsize=8)
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        out_path = os.path.join(output_dir, f"{prefix}_{metric}.png")
        plt.savefig(out_path, dpi=300, bbox_inches="tight")
        plt.close()
        print(f"  [Multiplot] Saved -> {out_path}")

def main():
    print("=" * 60)
    print("  GMR Fe3O4 — Classification Training")
    print("=" * 60)

    df = load_data(DATA_PATH, SHEET_NAME)
    print(f"  Dataset: {len(df)} samples | {df['concentration'].nunique()} classes")

    X = df[["delta_B_mT"]].values
    y = df["concentration"].values

    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=TEST_SIZE,
        random_state=RANDOM_STATE,
        stratify=y,
    )
    print(f"  Train: {len(X_train)} | Test: {len(X_test)}")

    models = build_models()
    skf = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)

    all_metrics = {}
    cv_results = {}
    report_dicts = {}
    cm_dict = {}
    preds = {}

    for name, model in models.items():
        print(f"\n  Training {name}...")
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        y_prob = model.predict_proba(X_test)

        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, average="macro", zero_division=0)
        rec = recall_score(y_test, y_pred, average="macro", zero_division=0)
        f1 = f1_score(y_test, y_pred, average="macro", zero_division=0)
        auc = roc_auc_score(y_test, y_prob, multi_class="ovr", average="macro")
        cv_sc = cross_val_score(model, X, y, cv=skf, scoring="accuracy")

        report = classification_report(
            y_test,
            y_pred,
            labels=CONCENTRATIONS,
            target_names=CLASS_NAMES,
            output_dict=True,
            zero_division=0,
        )
        cm = confusion_matrix(y_test, y_pred, labels=CONCENTRATIONS)

        all_metrics[name] = {
            "accuracy": float(acc),
            "precision": float(prec),
            "recall": float(rec),
            "f1": float(f1),
            "roc_auc": float(auc),
            "cv_mean": float(np.mean(cv_sc)),
            "cv_std": float(np.std(cv_sc)),
        }
        cv_results[name] = cv_sc
        report_dicts[name] = report
        cm_dict[name] = cm
        preds[name] = y_pred

        print(f"    Accuracy  : {acc:.4f}")
        print(f"    Precision : {prec:.4f} (macro)")
        print(f"    Recall    : {rec:.4f} (macro)")
        print(f"    F1        : {f1:.4f} (macro)")
        print(f"    ROC-AUC   : {auc:.4f} (OvR)")
        print(f"    CV Acc    : {np.mean(cv_sc):.4f} ± {np.std(cv_sc):.4f}")
        print(classification_report(
            y_test,
            y_pred,
            labels=CONCENTRATIONS,
            target_names=CLASS_NAMES,
            zero_division=0,
        ))

        model_dir = os.path.join(MODEL_DIR, name)
        ensure_dir(model_dir)
        model_path = os.path.join(model_dir, f"classifier_{name.lower()}.pkl")
        joblib.dump(model, model_path)
        print(f"    [Model] Saved -> {model_path}")

        out = dirs_for_classification_model(BASE_OUT_DIR, name)
        for key in ("excel", "json", "plots"):
            ensure_dir(out[key])

        plot_classification_individual(
            model_name=name,
            metrics=all_metrics[name],
            cm=cm_dict[name],
            y_test=y_test,
            y_pred=y_pred,
            report=report,
            plots_dir=out["plots"],
        )

        write_metrics_excel(
            {name: all_metrics[name]},
            {name: cv_results[name]},
            {name: report_dicts[name]},
            os.path.join(out["excel"], "classification_metrics.xlsx"),
        )
        save_metrics_json(
            {"model": name, "metrics": all_metrics[name], "classification_report": report_dicts[name]},
            os.path.join(out["json"], "classification_metrics.json"),
        )
        print(f"    [JSON] Saved -> {out['json']}")

    comp = dirs_classification_comparison(BASE_OUT_DIR)
    for key in ("excel", "json", "plots"):
        ensure_dir(comp[key])

    excel_path = os.path.join(comp["excel"], "classification_metrics.xlsx")
    write_metrics_excel(all_metrics, cv_results, report_dicts, excel_path)

    json_path = os.path.join(comp["json"], "classification_metrics.json")
    save_metrics_json(all_metrics, json_path)
    print(f"  [JSON] Saved -> {json_path}")

    print("\n  Generating comparison plot...")
    plot_classification_comparison(
        all_metrics=all_metrics,
        cv_results=cv_results,
        report_dicts=report_dicts,
        cm_dict=cm_dict,
        y_test=y_test,
        preds=preds,
        output_dir=comp["plots"],
    )

    print("\n  Generating metrics comparison plot...")
    plot_classification_metrics_comparison(all_metrics, comp["plots"])

    print("\n  Generating multi-metric comparison plots...")
    multiplot_root = os.path.join(comp["plots"], "multiplots")
    multiplot_4_dir = os.path.join(multiplot_root, "4_models")
    multiplot_3_dir = os.path.join(multiplot_root, "3_models")
    os.makedirs(multiplot_4_dir, exist_ok=True)
    os.makedirs(multiplot_3_dir, exist_ok=True)

    # 4 models multiplot
    plot_multi_metric_comparison_classification(
        all_metrics,
        list(all_metrics.keys()),
        [COLORS[n] for n in all_metrics.keys()],
        multiplot_4_dir,
        "multiplot_4_models.png",
        "Classification Models Comparison (4 Models)"
    )

    # 3 models separate per-metric plots (SVM, KNN, RandomForest)
    three_models = ["SVM", "KNN", "RandomForest"]
    three_metrics = {k: v for k, v in all_metrics.items() if k in three_models}
    three_colors = [COLORS[n] for n in three_models]
    plot_multi_metric_comparison_classification_separate(
        three_metrics,
        three_models,
        three_colors,
        multiplot_3_dir,
        "multiplot_3_models",
        "Classification Models Comparison (SVM, KNN, RF)"
    )

    print("\n" + "=" * 60)
    print("  Classification training complete.")
    print("=" * 60)

if __name__ == "__main__":
    main()